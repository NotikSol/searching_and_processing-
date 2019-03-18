import xlrd, xlwt
from openpyxl import load_workbook
import re
# убрать символы \xa0 и \n (['8:00', 'ИСТОРИЯ\xa0(пр)\nпроф.\xa0Дианов\xa0С.А. 413\xa0к.А\xa0(ЭТФ)'])
def xlsx_get_content(token = 'r.xlsx'):
  wb = load_workbook(token)
  sheet = wb['Лист1']
  row = 4
  i = row - 1
  table = []

  while row < 76:
    checkdata2 = sheet.cell(row, column=2).value
    checkdata3 = sheet.cell(row, column=3).value
    if checkdata3 != None: 
      temp = []
      if checkdata2 != None: 
        temp.append(checkdata2.replace("\n", " ").replace("\xa0", " ").strip())
        temp.append(checkdata3.replace("\n", " ").replace("\xa0", " ").strip())
        table.append(temp)
      else:
        temp.append(sheet.cell(i, column=2).value.replace("\n", " ").replace("\xa0", " ").strip())
        temp.append(checkdata3.replace("\n", " ").replace("\xa0", " ").strip())
        table.append(temp)
    row += 1
    i = row - 1
  #print(table)#test
  return table

if __name__ == '__main__':
  print(xlsx_get_content('r.xlsx'))